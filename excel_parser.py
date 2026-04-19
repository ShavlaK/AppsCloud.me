"""
Парсер Excel файлов прайс-листа (openpyxl).
Формат: 'Все товары.xlsx' — SKU, Категория, Название, Модель, Хранилище, Цена, Страна, Тип SIM
"""

import os
import logging
from typing import List, Dict, Optional
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelPriceParser:
    """Парсер Excel прайс-листа."""

    def __init__(
        self,
        column_name: str = "Название",
        column_price: str = "Цена",
        header_row: int = 1,
    ):
        """
        Args:
            column_name: Название колонки с названием товара
            column_price: Название колонки с ценой
            header_row: Номер строки заголовка (1-based)
        """
        self.column_name = column_name
        self.column_price = column_price
        self.header_row = header_row

    def parse_file(self, filepath: str) -> List[Dict[str, any]]:
        """
        Парсит Excel файл и возвращает список товаров с ценами.

        Args:
            filepath: Путь к .xlsx файлу

        Returns:
            [{"name": "Samsung A17 6/128 Gray", "price": 13000, "category": "Смартфоны"}, ...]
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Файл не найден: {filepath}")

        wb = load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active

        if ws is None:
            raise ValueError("Нет активного листа в Excel файле")

        # Находим индексы колонок по заголовку
        header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not header_cells:
            raise ValueError("Заголовки не найдены")

        headers = [str(h).strip() for h in header_cells[0]]
        logger.info(f"Найдены колонки: {headers}")

        try:
            name_idx = headers.index(self.column_name)
            price_idx = headers.index(self.column_price)
        except ValueError as e:
            raise ValueError(f"Колонка не найдена в заголовках: {e}") from e

        # Ищем дополнительные колонки
        category_idx = None
        try:
            category_idx = headers.index("Категория")
        except ValueError:
            pass

        # Парсим строки данных
        prices = []
        row_count = 0
        skipped = 0

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            row_count += 1

            name = row[name_idx] if name_idx < len(row) else None
            price = row[price_idx] if price_idx < len(row) else None

            if not name or name == "":
                skipped += 1
                continue

            try:
                price_val = float(str(price).replace(" ", "").replace(",", ""))
            except (ValueError, TypeError):
                skipped += 1
                logger.warning(f"Строка {row_idx}: не удалось распознать цену '{price}'")
                continue

            item: Dict[str, any] = {
                "name": str(name).strip(),
                "price": int(price_val),
                "original_price": int(price_val),
            }

            if category_idx is not None and category_idx < len(row):
                cat = row[category_idx]
                if cat:
                    item["category"] = str(cat).strip()

            prices.append(item)

        wb.close()

        logger.info(
            f"Спарсено {len(prices)} позиций из {row_count} строк "
            f"(пропущено: {skipped})"
        )

        return prices

    def apply_markup(
        self, prices: List[Dict[str, any]], markup: float
    ) -> List[Dict[str, any]]:
        """
        Применяет наценку к ценам.

        Args:
            prices: Список товаров
            markup: Сумма наценки (если < 100 — процент)

        Returns:
            Обновлённый список
        """
        updated = []
        for item in prices:
            original = item["price"]

            if markup < 100:
                new_price = int(original * (1 + markup / 100))
            else:
                new_price = original + markup

            updated.append({
                "name": item["name"],
                "original_price": original,
                "price": new_price,
                "category": item.get("category", ""),
            })

        return updated
